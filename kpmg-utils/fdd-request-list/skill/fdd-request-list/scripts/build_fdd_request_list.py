#!/usr/bin/env python3
"""Build a KPMG-style FDD request-list workbook from a JSON spec."""

from __future__ import annotations

import argparse
import json
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_NAME = "Financial Requests"
BODY_START_ROW = 23
PRIORITY_TEMPLATE_ROW = 23
CONDITIONAL_TEMPLATE_ROW = 48
ITEM_TEMPLATE_ROW = 24
PROJECT_TITLE_TEMPLATE = "Project {project_name} - Financial Information Request List"
PRIORITY_TITLES = {
    "highest priority",
    "high priority",
    "medium priority",
    "low priority",
}
CONTACT_FIELD_ROWS = {
    "manager_name": "B15",
    "firm_name": "B16",
    "email": "B17",
    "phone": "B18",
    "cell": "B19",
}
TOP_LEVEL_CELL_MAP = {
    "historical_period_text": "B11",
}


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Build an FDD request-list workbook from a JSON spec.",
    )
    parser.add_argument("--spec", required=True, help="Path to the input JSON spec.")
    parser.add_argument("--output", required=True, help="Path to the output .xlsx workbook.")
    parser.add_argument(
        "--template",
        help="Optional workbook template override. Defaults to the bundled asset.",
    )
    return parser.parse_args()


def bundled_template_path() -> Path:
    """Return the default bundled workbook template path."""
    return Path(__file__).resolve().parents[1] / "assets" / "example-information-request-list-template.xlsx"


def load_spec(spec_path: Path) -> dict[str, Any]:
    """Load and lightly validate the JSON spec."""
    spec = json.loads(spec_path.read_text())
    if not isinstance(spec, dict):
        raise ValueError("Spec root must be a JSON object.")
    sections = spec.get("sections")
    if not isinstance(sections, list) or not sections:
        raise ValueError("Spec must include a non-empty 'sections' array.")
    for section in sections:
        if not isinstance(section, dict):
            raise ValueError("Each section must be an object.")
        if not isinstance(section.get("title"), str) or not section["title"].strip():
            raise ValueError("Each section must include a non-empty string title.")
        items = section.get("items", [])
        if items is None:
            section["items"] = []
        elif not isinstance(items, list):
            raise ValueError(f"Section '{section['title']}' has a non-list 'items' value.")
    return spec


def parse_excel_date(value: Any) -> date | datetime | None:
    """Parse ISO-like date input into a Python date or datetime."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return date.fromisoformat(text)
    raise ValueError(f"Unsupported date value: {value!r}")


def clone_row_format(template_ws, template_row: int, output_ws, output_row: int) -> None:
    """Clone row formatting from the template into the output worksheet."""
    # Copy the visible row framing across the full A:J span.
    for column_index in range(1, 11):
        source_cell = template_ws.cell(template_row, column_index)
        target_cell = output_ws.cell(output_row, column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    source_dimensions = template_ws.row_dimensions[template_row]
    target_dimensions = output_ws.row_dimensions[output_row]
    target_dimensions.height = source_dimensions.height
    target_dimensions.hidden = source_dimensions.hidden


def clear_row_values(ws, row_index: int) -> None:
    """Clear values for the full visible row span without disturbing formatting."""
    for column_index in range(1, 11):
        ws.cell(row_index, column_index).value = None


def is_priority_section(title: str) -> bool:
    """Return True when the section title should use the priority style."""
    return title.strip().lower() in PRIORITY_TITLES


def update_header_cells(ws, spec: dict[str, Any]) -> None:
    """Populate the non-body template fields from the spec."""
    for spec_key, cell_ref in TOP_LEVEL_CELL_MAP.items():
        value = spec.get(spec_key)
        if value:
            ws[cell_ref] = value

    project_name = str(spec.get("project_name") or "[NAME]").strip() or "[NAME]"
    ws["B21"] = PROJECT_TITLE_TEMPLATE.format(project_name=project_name)

    contact_fields = spec.get("contact_fields", {})
    if not isinstance(contact_fields, dict):
        raise ValueError("'contact_fields' must be an object when provided.")
    for field_name, cell_ref in CONTACT_FIELD_ROWS.items():
        value = contact_fields.get(field_name)
        if value not in (None, ""):
            ws[cell_ref] = value


def build_body_rows(spec: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten sections and items into a renderable row sequence."""
    rows: list[dict[str, Any]] = []
    fallback_id = 1

    for section in spec["sections"]:
        rows.append(
            {
                "row_type": "priority_section" if is_priority_section(section["title"]) else "conditional_section",
                "title": section["title"],
            }
        )
        for item in section.get("items", []):
            if not isinstance(item, dict):
                raise ValueError(f"Section '{section['title']}' contains a non-object item.")
            request_id = item.get("request_id")
            if request_id in (None, ""):
                request_id = fallback_id
            rows.append(
                {
                    "row_type": "item",
                    "request_id": request_id,
                    "request_text": item.get("request_text", ""),
                    "priority": item.get("priority", ""),
                    "dataroom_reference": item.get("dataroom_reference", ""),
                    "request_date": item.get("request_date", spec.get("request_date")),
                    "status": item.get("status") or "Open",
                    "kpmg_comment": item.get("kpmg_comment", ""),
                    "target_response": item.get("target_response", ""),
                }
            )
            fallback_id += 1

    return rows


def render_body(template_ws, output_ws, body_rows: list[dict[str, Any]]) -> None:
    """Replace the template body with the generated request rows."""
    existing_body_rows = max(output_ws.max_row - BODY_START_ROW + 1, 0)
    if existing_body_rows:
        output_ws.delete_rows(BODY_START_ROW, existing_body_rows)

    if not body_rows:
        return

    output_ws.insert_rows(BODY_START_ROW, amount=len(body_rows))

    for offset, row_data in enumerate(body_rows):
        row_index = BODY_START_ROW + offset
        clear_row_values(output_ws, row_index)

        row_type = row_data["row_type"]
        if row_type == "priority_section":
            clone_row_format(template_ws, PRIORITY_TEMPLATE_ROW, output_ws, row_index)
            output_ws[f"B{row_index}"] = row_data["title"]
            continue

        if row_type == "conditional_section":
            clone_row_format(template_ws, CONDITIONAL_TEMPLATE_ROW, output_ws, row_index)
            output_ws[f"B{row_index}"] = row_data["title"]
            continue

        clone_row_format(template_ws, ITEM_TEMPLATE_ROW, output_ws, row_index)
        output_ws[f"B{row_index}"] = row_data["request_id"]
        output_ws[f"C{row_index}"] = row_data["request_text"]
        output_ws[f"D{row_index}"] = row_data["priority"]
        output_ws[f"E{row_index}"] = row_data["dataroom_reference"]
        output_ws[f"F{row_index}"] = parse_excel_date(row_data["request_date"])
        output_ws[f"G{row_index}"] = row_data["status"]
        output_ws[f"H{row_index}"] = row_data["kpmg_comment"]
        output_ws[f"I{row_index}"] = row_data["target_response"]


def build_workbook(spec: dict[str, Any], template_path: Path, output_path: Path) -> Path:
    """Build the output workbook from the template and spec."""
    template_wb = load_workbook(template_path)
    output_wb = load_workbook(template_path)
    template_ws = template_wb[SHEET_NAME]
    output_ws = output_wb[SHEET_NAME]

    update_header_cells(output_ws, spec)
    body_rows = build_body_rows(spec)
    render_body(template_ws, output_ws, body_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)
    return output_path


def main() -> None:
    """Run the CLI entrypoint."""
    args = parse_args()
    spec_path = Path(args.spec).resolve()
    output_path = Path(args.output).resolve()
    template_path = Path(args.template).resolve() if args.template else bundled_template_path()

    spec = load_spec(spec_path)
    result_path = build_workbook(spec, template_path, output_path)
    section_count = len(spec["sections"])
    item_count = sum(len(section.get("items", [])) for section in spec["sections"])
    print(f"Wrote {result_path} ({section_count} sections, {item_count} requests)")


if __name__ == "__main__":
    main()
