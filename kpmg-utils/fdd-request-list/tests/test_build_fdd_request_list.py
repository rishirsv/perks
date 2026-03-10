"""Tests for the FDD request-list workbook generator."""

from __future__ import annotations

import json
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SKILL_ROOT = ROOT / "skill" / "fdd-request-list"
SCRIPT_PATH = SKILL_ROOT / "scripts" / "build_fdd_request_list.py"
TEMPLATE_PATH = SKILL_ROOT / "assets" / "example-information-request-list-template.xlsx"
FIXTURES_DIR = ROOT / "tests" / "fixtures"


def run_builder(spec_name: str) -> tuple[Path, tempfile.TemporaryDirectory[str]]:
    """Run the workbook generator for a fixture and return the output path and tempdir."""
    spec_path = FIXTURES_DIR / spec_name
    temp_dir = tempfile.TemporaryDirectory()
    output_path = Path(temp_dir.name) / spec_path.with_suffix(".xlsx").name
    command = [
        "python3",
        str(SCRIPT_PATH),
        "--spec",
        str(spec_path),
        "--output",
        str(output_path),
    ]
    subprocess.run(command, check=True, capture_output=True, text=True)
    return output_path, temp_dir


class BuildFddRequestListTests(unittest.TestCase):
    """Verify workbook generation and template fidelity."""

    def setUp(self) -> None:
        self.template_wb = load_workbook(TEMPLATE_PATH)
        self.template_ws = self.template_wb["Financial Requests"]

    def test_cli_generates_all_scenarios(self) -> None:
        """Generate representative scenario workbooks successfully."""
        scenarios = [
            "core-fdd.json",
            "carve-out-audited.json",
            "software-subscription.json",
            "energy-regulated-utility.json",
        ]

        for spec_name in scenarios:
            with self.subTest(spec_name=spec_name):
                output_path, temp_dir = run_builder(spec_name)
                self.addCleanup(temp_dir.cleanup)
                workbook = load_workbook(output_path)
                worksheet = workbook["Financial Requests"]

                spec = json.loads((FIXTURES_DIR / spec_name).read_text())
                self.assertEqual(worksheet.title, "Financial Requests")
                self.assertEqual(
                    [str(item) for item in worksheet.merged_cells.ranges],
                    [str(item) for item in self.template_ws.merged_cells.ranges],
                )
                self.assertEqual(
                    worksheet["B21"].value,
                    f"Project {spec['project_name']} - Financial Information Request List",
                )
                self.assertEqual(worksheet["B23"].value, spec["sections"][0]["title"])

    def test_header_styles_and_contract_are_preserved(self) -> None:
        """Keep the template framing intact after generation."""
        output_path, temp_dir = run_builder("core-fdd.json")
        self.addCleanup(temp_dir.cleanup)
        worksheet = load_workbook(output_path)["Financial Requests"]

        for cell_ref in ("B5", "B7", "B11", "B21", "B22", "C22", "H22"):
            with self.subTest(cell_ref=cell_ref):
                self.assertEqual(worksheet[cell_ref]._style, self.template_ws[cell_ref]._style)

        self.assertEqual(worksheet["B22"].value, "#")
        self.assertEqual(worksheet["C22"].value, "Request")

    def test_section_and_item_styles_match_template_rows(self) -> None:
        """Apply the correct template row styles to generated content."""
        output_path, temp_dir = run_builder("carve-out-audited.json")
        self.addCleanup(temp_dir.cleanup)
        worksheet = load_workbook(output_path)["Financial Requests"]

        # First priority section row.
        self.assertEqual(worksheet["B23"]._style, self.template_ws["B23"]._style)
        # First request item row.
        self.assertEqual(worksheet["C24"]._style, self.template_ws["C24"]._style)
        # First conditional section row after the initial highest-priority items.
        self.assertEqual(worksheet["B26"].value, "If audited (and workpaper access necessary / reconciliation to the audit not straight-forward)")
        self.assertEqual(worksheet["B26"]._style, self.template_ws["B48"]._style)

    def test_request_dates_and_values_are_written(self) -> None:
        """Populate item values, statuses, and Excel dates correctly."""
        output_path, temp_dir = run_builder("software-subscription.json")
        self.addCleanup(temp_dir.cleanup)
        worksheet = load_workbook(output_path)["Financial Requests"]

        self.assertEqual(worksheet["B24"].value, 1)
        self.assertEqual(worksheet["D24"].value, "Highest")
        self.assertEqual(worksheet["G24"].value, "Open")
        self.assertEqual(worksheet["H24"].value, "Use this to bridge recurring revenue by stream.")
        self.assertEqual(str(worksheet["F24"].value.date()), "2026-03-10")


if __name__ == "__main__":
    unittest.main()
